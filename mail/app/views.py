from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import *
from .forms import *
from django.contrib import messages
from django.shortcuts import redirect
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse

class EmailSettingsListView(ListView):
    model = EmailSettings
    template_name = 'settings_list.html'
    context_object_name = 'settings'

class EmailSettingsCreateView(CreateView):
    model = EmailSettings
    form_class = EmailSettingsForm
    template_name = 'settings_form.html'
    success_url = reverse_lazy('settings_list')

class EmailSettingsUpdateView(UpdateView):
    model = EmailSettings
    form_class = EmailSettingsForm
    template_name = 'settings_form.html'
    success_url = reverse_lazy('settings_list')

class EmailGroupListView(ListView):
    model = EmailGroup
    template_name = 'group_list.html'
    context_object_name = 'groups'

class EmailGroupCreateView(CreateView):
    model = EmailGroup
    form_class = EmailGroupForm
    template_name = 'group_form.html'
    success_url = reverse_lazy('group_list')

class EmailGroupUpdateView(UpdateView):
    model = EmailGroup
    form_class = EmailGroupForm
    template_name = 'group_form.html'
    success_url = reverse_lazy('group_list')

class EmailGroupDeleteView(DeleteView):
    model = EmailGroup
    template_name = 'group_confirm_delete.html'
    success_url = reverse_lazy('group_list')

class EmailsListView(ListView):
    model = Emails
    template_name = 'email_list.html'
    context_object_name = 'emails'

    def get_queryset(self):
        # Get the base queryset
        queryset = super().get_queryset()
        
        # Get the group filter from the query string (e.g., ?group=1)
        group_id = self.request.GET.get('group')
        
        # Apply filter if group_id is provided and valid
        if group_id:
            try:
                queryset = queryset.filter(email_group__id=group_id)
            except ValueError:
                # Handle invalid group_id (e.g., non-integer)
                pass
        
        return queryset

    def get_context_data(self, **kwargs):
        # Add all groups to context for the filter dropdown
        context = super().get_context_data(**kwargs)
        context['groups'] = EmailGroup.objects.all()
        context['selected_group'] = self.request.GET.get('group')
        return context

class EmailsCreateView(CreateView):
    model = Emails
    form_class = EmailsForm
    template_name = 'email_form.html'
    success_url = reverse_lazy('email_list')

class EmailsUpdateView(UpdateView):
    model = Emails
    form_class = EmailsForm
    template_name = 'email_form.html'
    success_url = reverse_lazy('email_list')

class EmailsDeleteView(DeleteView):
    model = Emails
    template_name = 'email_confirm_delete.html'
    success_url = reverse_lazy('email_list')


class EmailMessageListView( ListView):
    model = EmailMessage
    template_name = 'message_list.html'
    context_object_name = 'messages'

class EmailMessageCreateView( CreateView):
    model = EmailMessage
    form_class = EmailMessageForm
    template_name = 'message_form.html'
    success_url = reverse_lazy('message_list')
    permission_required = 'emails.add_emailmessage'
    
    def form_valid(self, form):
        # Save the form first
        response = super().form_valid(form)
        
        # Send email immediately after creation
        try:
            self.object.send_email()
            messages.success(self.request, f"Email '{self.object.subject}' sent successfully!")
        except Exception as e:
            messages.error(self.request, f"Failed to send email: {str(e)}")
            
        return response

class EmailMessageUpdateView(UpdateView):
    model = EmailMessage
    form_class = EmailMessageForm
    template_name = 'message_form.html'
    success_url = reverse_lazy('message_list')
    permission_required = 'emails.change_emailmessage'

class EmailMessageDeleteView(DeleteView):
    model = EmailMessage
    template_name = 'message_confirm_delete.html'
    success_url = reverse_lazy('message_list')
    permission_required = 'emails.delete_emailmessage'

def send_email_message(request, pk):
    try:
        message = EmailMessage.objects.get(pk=pk)
        print(f"Found message: {message.subject}")  # Debug: Confirm message exists
    except EmailMessage.DoesNotExist:
        messages.error(request, f"No email message found with ID {pk}")
        print(f"Error: No message with pk={pk}")  # Debug output
        return redirect('message_list')

    try:
        print("Attempting to send email...")  # Debug: Before sending
        message.send_email()
        messages.success(request, f"Email '{message.subject}' sent successfully!")
        print("Email sent successfully")  # Debug: After success
    except Exception as e:
        error_msg = f"Failed to send email: {str(e)}"
        messages.error(request, error_msg)
        print(error_msg)  # Debug: Print exact error
    return redirect('message_list')




def export_emails(request):
    group_id = request.GET.get('group')
    if not group_id:
        messages.error(request, "Please select a group to export.")
        return redirect('email_list')

    try:
        group = EmailGroup.objects.get(id=group_id)
        emails = Emails.objects.filter(email_group=group)

        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = f"{group.name} Emails"
        ws.append(['Email'])  # Header

        for email in emails:
            ws.append([email.email])

        # Prepare response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{group.name}_emails.xlsx"'
        wb.save(response)
        return response

    except EmailGroup.DoesNotExist:
        messages.error(request, "Selected group does not exist.")
        return redirect('email_list')

def import_emails(request):
    if request.method == 'POST':
        print("POST request received")  # Debug: Confirm POST
        group_id = request.POST.get('group')
        file = request.FILES.get('excel_file')
        
        print(f"Group ID: {group_id}")  # Debug: Check group selection
        print(f"File: {file}")  # Debug: Check file upload

        if not group_id or not file:
            messages.error(request, "Please select a group and upload a file.")
            print("Missing group or file")  # Debug
            return redirect('email_list')

        try:
            group = EmailGroup.objects.get(id=group_id)
            print(f"Group found: {group.name}")  # Debug: Confirm group
            
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            print(f"Worksheet loaded: {ws.title}")  # Debug: Confirm file read

            # Check header
            header = [cell.value for cell in ws[1]]
            print(f"Header: {header}")  # Debug: Verify header
            if not header or header[0] != 'Email':
                messages.error(request, "Invalid file format. Expected header: 'Email'")
                print("Invalid header")  # Debug
                return redirect('email_list')

            # Process rows
            emails_added = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                print(f"Row data: {row}")  # Debug: Show each row
                email = row[0]
                if email and isinstance(email, str):
                    email = email.strip()
                    print(f"Processing email: {email}")  # Debug
                    email_obj, created = Emails.objects.get_or_create(
                        email_group=group,
                        email=email
                    )
                    if created:
                        emails_added += 1
                        print(f"Added email: {email}")  # Debug
                    else:
                        print(f"Email already exists: {email}")  # Debug

            messages.success(request, f"Successfully imported {emails_added} emails.")
            print(f"Imported {emails_added} emails")  # Debug
        except EmailGroup.DoesNotExist:
            messages.error(request, "Selected group does not exist.")
            print("Group not found")  # Debug
        except Exception as e:
            messages.error(request, f"Error importing emails: {str(e)}")
            print(f"Error: {str(e)}")  # Debug

        return redirect('email_list')

    print("Not a POST request")  # Debug: Catch non-POST
    return redirect('email_list')